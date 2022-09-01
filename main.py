#Program tu get the values from the txt files and write them in to the excel sheet 
#have to put the script in the same directory than the txt file and excel sheet

## Ricardo Caldera   8/25/2022


####### IMPORT SECTION
import sys
import openpyxl
import string
from openpyxl.descriptors import base
from colorama import init
init()
from colorama import Fore,Back,Style

###### CONSTANT SECTION 
SUBSTRING_START = "==============="
SUBSTRING_END = "Est. SPECrate"
ASTERISK = "*"
TIME = '-t'

#### CLASS SECTION
class Validator():
    def __init__(self,argv=False) -> None:
        self.argv = argv
    
    def validate(self):
        if not self.validateParams():
            return False
        if not self.checkParamsOrder():
            return False
        return True
    
    def validateParams(self):
        if len(self.argv) <4:
            self.missingParamsMessage()
            return False
        return True

    def missingParamsMessage(self):
        self.printErrorMessage("Missing parameters")
        
    def wrongOrderMessage(self):
        self.printErrorMessage("Invalid order")

    def printErrorMessage(self,text):
        print("\n")
        print(Fore.RED + text)
        print(Fore.BLUE + "main.exe <file.txt> <file.xlsx> <cell> <time>")

    def checkParamsOrder(self):
        txtFile = str(self.argv[1])
        excelFile = str(self.argv[2])
        time = str(sys.argv[4]).lower() == TIME if len(sys.argv) > 4 else False


        if len(self.argv[3])<2:
            self.printErrorMessage("wrong cell")
            return False

        letter = str(self.argv[3][0])
        number = int(self.argv[3][1:])

        if txtFile.split(".")[-1] != "txt":
            self.printErrorMessage("wrong txt file")
            return False
        if excelFile.split(".")[-1] != "xlsx":
            self.printErrorMessage("wrong xlsx file")
            return False
        if not letter.isalpha():
            self.printErrorMessage("wrong cell")
            return False
        if number<1:
            self.printErrorMessage("wrong cell")
            return False
        if time and letter.lower() == "a":
            self.printErrorMessage("wrong cell")
            return False
        return True


### INIT SECTION
validator = Validator(sys.argv)

        
###### ASSIGN SECTION
if not validator.validate():
    exit()


TEXT_FILE = str(sys.argv[1])
EXCEL_FILE = str(sys.argv[2])
CELL = str(sys.argv[3])
LETTER = CELL[0]
ADD_BASE_RUN_TIME = str(sys.argv[4]).lower() == TIME if len(sys.argv) > 4 else False
WORK_BOOK = openpyxl.load_workbook(EXCEL_FILE)
SHEET = WORK_BOOK.active
column = string.ascii_lowercase.index(LETTER.lower()) +1 
row = int(CELL[1:])
baseRunTime = []
baseRate = []

###### FUNCTIONS SECTION

    


def checkSaveValues(line,pValue):
    if SUBSTRING_START in line:
        return True,False
    if SUBSTRING_END in line:
        return False,True
    return pValue,False

def getRunTimeAndRate(line):
    array = line.split()
    if len(array) == 5:
        baseRunTime.append(int(array[2])) 
        baseRate.append(float(array[3])) 
        return
    elif len(array)==3:
        baseRate.append(float(array[2])) 
        return



def getValuesFromTxt(textFile):
    saveValues = False
    done = False
    with open(textFile,"r") as file:
        for i,line in enumerate(file):
            if saveValues:
                getRunTimeAndRate(line)
            saveValues,done = checkSaveValues(line,saveValues)
            if done:
                break

def setBaseRateValues(row,column):
    for rate in baseRate:
        SHEET.cell(row,column,rate)
        row = row+1
    return

def setBaseRunTimeValues(row,column):
    for rate in baseRunTime:
       SHEET.cell(row,column-1,rate)
       row = row+1 

def setValuesToExcel(row,column):
    setBaseRateValues(row,column)
    if ADD_BASE_RUN_TIME:
        setBaseRunTimeValues(row,column) 
    WORK_BOOK.save(EXCEL_FILE)

def doneMessage():
    print("\n")
    print(Fore.GREEN+"************************************")
    print(Fore.GREEN+"Data imported successfully")
    print(Fore.GREEN+"************************************")
def wrongOrderMessage():
    print("\n")
    print(Fore.RED + "Invalid order")
    print(Fore.BLUE + "main.exe <file.txt> <file.xlsx> <cell> <time>")



###### MAIN SECTION

getValuesFromTxt(TEXT_FILE)
setValuesToExcel(row,column)
doneMessage()



    
    